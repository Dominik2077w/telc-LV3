#!/usr/bin/env python3
"""Extract LV2 paragraph flashcard practice data from the cleaned workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE_SHEET = "LV2清洗数据"
SKIP_TEILS = {"1"}
UNUSABLE_MARKERS = ("资料未明确", "不作为配题答案", "未配题")

TEXT_REPLACEMENTS = (
    ("ZumGlück", "Zum Glück"),
    ("AbgewissenGrößenordnungensindSinnfragenangebracht.", "Ab gewissen Größenordnungen sind Sinnfragen angebracht."),
    ("Esist Zeit", "Es ist Zeit"),
    ("Nein,Abschlussarbeiten", "Nein, Abschlussarbeiten"),
    ("DochwelcheBedeutunghabenTräumeeigentlich?", "Doch welche Bedeutung haben Träume eigentlich?"),
    (
        "DieseSchwierigkeitkönnteauchdarinbegründet liegen,dassderGegenstandderUntersuchung nurschwerzufassenist.",
        "Diese Schwierigkeit könnte auch darin begründet liegen, dass der Gegenstand der Untersuchung nur schwer zu fassen ist.",
    ),
    (
        "EineinteressanteSpurverfolgendabeiNeuroforrscher,diebildgebendeVerfahrennutzen,umdas träumendeGehirnzubeobachten.",
        "Eine interessante Spur verfolgen dabei Neuroforscher, die bildgebende Verfahren nutzen, um das träumende Gehirn zu beobachten.",
    ),
    (
        "Währendich hustendauf dem Sofaliege und übermeinLeben nachdenke,habe ich diesenWerbespotim Kopf",
        "Während ich hustend auf dem Sofa liege und über mein Leben nachdenke, habe ich diesen Werbespot im Kopf",
    ),
    (
        "SoskeptischWissenschaftlerdasPhänomenderMondglaubigkeitauchsehenmögen-das InteresseandemHimmelskörperistnachwievorsehrgroß.",
        "So skeptisch Wissenschaftler das Phänomen der Mondgläubigkeit auch sehen mögen - das Interesse an dem Himmelskörper ist nach wie vor sehr groß.",
    ),
    (
        "NeulichsaßichnachmittagsanderUniineinerVorlesungundhörteplötzlich,wiemeinPhonevibrierte.",
        "Neulich saß ich nachmittags an der Uni in einer Vorlesung und hörte plötzlich, wie mein iPhone vibrierte.",
    ),
    (
        "DochwarumistmeineAbneigunggegendasTelefonierensostark?",
        "Doch warum ist meine Abneigung gegen das Telefonieren so stark?",
    ),
    (
        "WennichüberdasheutigeKommunikationsverhaltennachdenke,fälltmirallerdingseinmerkwürdiger Widerspruchauf:Istesniechterstaunlich,dassalleständigeinGerätmitsichherumtrgen,dasursprünglichin ersterLiniezumTelefonierengedachtwar-abergeradedieseFunktionimmerwenigergenutztwird?",
        "Wenn ich über das heutige Kommunikationsverhalten nachdenke, fällt mir allerdings ein merkwürdiger Widerspruch auf: Ist es nicht erstaunlich, dass alle ständig ein Gerät mit sich herumtragen, das ursprünglich in erster Linie zum Telefonieren gedacht war - aber gerade diese Funktion immer weniger genutzt wird?",
    ),
    (
        "MitmeinerAbnerigunggegendasTelefonierensteheichübrigensnichtalleineda.",
        "Mit meiner Abneigung gegen das Telefonieren stehe ich übrigens nicht allein da.",
    ),
    ("dieVerkehrssicherheit", "die Verkehrssicherheit"),
    ("durch aus", "durchaus"),
    ("Ge bäckstücks", "Gebäckstücks"),
    ("Schwer mütigen", "Schwermütigen"),
    ("Umstel lung", "Umstellung"),
    ("zufriedenzu stellen", "zufrieden zu stellen"),
    ("visuellen Welten des Internets und der sozialen Meaien", "visuellen Welten des Internets und der sozialen Medien"),
    ("ohneden Rückzug", "ohne den Rückzug"),
    ("vielenandern", "vielen anderen"),
    ("Diese Zeiten sind zu, Glück", "Diese Zeiten sind zum Glück"),
    ("ist des Fernstudiums", "ist die Stärke des Fernstudiums"),
    ("gleichzeitig sein größte Nachteil", "gleichzeitig sein größter Nachteil"),
    ("Überfordert", "überfordert"),
    ("Kunde Studienanbieter", "Kunden der Studienanbieter"),
    (" GIücksempfinden ", " Glücksempfinden "),
    ("Dinosaurier:", "Dinosauriern:"),
    ("Kevin Weyerjedenfalls", "Kevin Weyer jedenfalls"),
    ("Master Weiterbildungsstudium", "Master-Weiterbildungsstudium"),
    ("diverser-teilweise", "diverser - teilweise"),
    ("Vorn am Pult stand eine allwissende Respektsperson die", "Vorn am Pult stand eine allwissende Respektsperson, die"),
    ("Schwerkrafteffekt", "Schwerkrafteffekt"),
    ("nicht mehr gefragt war Sicher ist", "nicht mehr gefragt war. Sicher ist"),
    ("Was in den Texten zu dem Thema nicht mal am Rande erwähnt wird, Es gibt", "Was in den Texten zu dem Thema nicht mal am Rande erwähnt wird: Es gibt"),
    ("und Es gibticht, wed sie es mussten", "und es gibt sie nicht, weil sie es müssen"),
    ("allein ins Restaurant geht", "allein ins Restaurant gehen"),
    ("Doch wie kann Konsum überhaupt Umweltbewusstsein?", "Doch wie kann Konsum überhaupt Umweltbewusstsein ausdrücken?"),
    ("Käufe–", "Käufe -"),
    ("nichts einzuwenden–", "nichts einzuwenden -"),
    ("exzessivem Konsum", "exzessivem Konsum"),
    ("Man konnte so beginnen: die folgenden Untersuchung", "Man könnte so beginnen: Die folgende Untersuchung"),
    ("Von Rollendtrukturen", "von Rollenstrukturen"),
    ("minimalen Organisation", "minimalen Organisationen"),
    ("amerikanische Soziologie", "amerikanische Soziologe"),
    ("Disziplin, die", "Disziplinen, die"),
    ("wer etwa", "Wer etwa"),
    ("nur Solche", "nur solche"),
    ("Eltern können anstrengenden sein.", "Eltern können anstrengend sein."),
    ("wer hat sich nicht schon mal selbstdabei", "wer hat sich nicht schon mal selbst dabei"),
    ("Hand aufs Herz,wer", "Hand aufs Herz, wer"),
    ("Beispiele gefällig etwa", "Beispiele gefällig? Etwa"),
    ("Verschieben Sie die Deutscharbeit-", "Verschieben Sie die Deutscharbeit -"),
    ("wann der erste Elternabend", "Wann der erste Elternabend"),
    ("Eine coole Rabenmutter Auch", "Eine coole Rabenmutter? Auch"),
    ("räumenim", "räumen im"),
    ("Ge bäckstücks", "Gebäckstücks"),
)

PROMPT_REPLACEMENTS = (
    ("äußertderAutorBedenken?", "äußert der Autor Bedenken?"),
    ("LiefertderAutorreineFakten?", "Liefert der Autor reine Fakten?"),
    ("Amüsieren sich der Autor?", "Amüsiert sich der Autor?"),
    ("äußertdieAutorinBedenken?wirftdieAutorinBedenkenauf.? wirftdieAutorinZweifelauf.? äußertsichdieAutorinabfällig?", "äußert die Autorin Bedenken? Wirft die Autorin Zweifel auf? Äußert sich die Autorin abfällig?"),
    ("stelltdieAutorinverschiedeneStandpunkteeinanderGegenüber?", "stellt die Autorin verschiedene Standpunkte einander gegenüber?"),
    ("ziehtdieAutorineineSchlussfolgerung?", "zieht die Autorin eine Schlussfolgerung?"),
    ("liefertdieAutorinDefinition?", "liefert die Autorin eine Definition?"),
    ("äußertderAutoreinenWunsch?", "äußert der Autor einen Wunsch?"),
    ("gibtdieAutorineinenRat?", "gibt die Autorin einen Rat?"),
    ("möchtederAutorzumNachdenkenanregen?", "möchte der Autor zum Nachdenken anregen?"),
    ("suchtdieAutorinbeidenLesernnachZustimmung?", "sucht die Autorin bei den Lesern nach Zustimmung?"),
    ("bringtdieAutorinVergleichean?", "bringt die Autorin Vergleiche an?"),
    ("drücktdieAutorinVerwunderungaus?", "drückt die Autorin Verwunderung aus?"),
    ("lädtdieAutorindieLeserzueinemgedanklichenExperimentein?", "lädt die Autorin die Leser zu einem gedanklichen Experiment ein?"),
    ("beruft sich die Autirin auf die Eindvhätzungen von Fachleuten?", "beruft sich die Autorin auf die Einschätzungen von Fachleuten?"),
    ("möchtedie Autorin warnen?", "möchte die Autorin warnen?"),
    ("wagt die Autorin eine Prognose? A", "wagt die Autorin eine Prognose?"),
)


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_ocr_text(value: str, *, prompt: bool = False) -> str:
    text = clean_text(value)
    replacements = (*TEXT_REPLACEMENTS, *(PROMPT_REPLACEMENTS if prompt else ()))
    for old, new in replacements:
        text = text.replace(old, new)
    text = re.sub(r"\s+([,.;:?!])", r"\1", text)
    text = re.sub(r"([,;:])(?=\S)", r"\1 ", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()


def split_answer_items(numbers: str, item_text: str) -> list[tuple[str, str]]:
    text = clean_text(item_text)
    if not text or any(marker in text for marker in UNUSABLE_MARKERS):
        return []

    number_parts = [part.strip() for part in re.split(r"[;；,，]", clean_text(numbers)) if part.strip()]
    raw_parts = [part.strip() for part in re.split(r"[;；]", text) if part.strip()]

    if len(raw_parts) == 1:
        matches = list(re.finditer(r"(?<!\d)(\d{1,2})[\s.\-]+", raw_parts[0]))
        if len(matches) > 1:
            expanded = []
            for index, match in enumerate(matches):
                start = match.start()
                end = matches[index + 1].start() if index + 1 < len(matches) else len(raw_parts[0])
                expanded.append(raw_parts[0][start:end].strip())
            raw_parts = expanded

    results: list[tuple[str, str]] = []
    for index, part in enumerate(raw_parts):
        match = re.match(r"^\s*(\d{1,2})\s*[-.)]?\s*(.*)$", part)
        if match:
            number, prompt = match.group(1), match.group(2)
        else:
            number = number_parts[index] if index < len(number_parts) else ""
            prompt = part

        prompt = clean_ocr_text(prompt, prompt=True)
        prompt = re.sub(r"^[A-E]\s*\)\s*\(?\s*", "", prompt).strip()
        prompt = prompt.strip("() ")
        if not prompt:
            continue
        results.append((number, prompt))

    return results


def build_sections(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[SOURCE_SHEET]
    rows = list(worksheet.iter_rows(values_only=True))
    header = [clean_text(cell) for cell in rows[0]]
    indexes = {name: index for index, name in enumerate(header)}

    sections: dict[str, dict] = {}

    for row in rows[1:]:
        teil = clean_text(row[indexes["Teil编号"]])
        if not teil or teil in SKIP_TEILS:
            continue

        title = clean_ocr_text(row[indexes["篇目标题"]])
        paragraph = clean_text(row[indexes["段落编号"]])
        first_sentence = clean_ocr_text(row[indexes["第一句话原文"]])
        summary = clean_ocr_text(row[indexes["段落大意（非逐句译文）"]])
        numbers = clean_text(row[indexes["对应题号"]])
        item_text = clean_ocr_text(row[indexes["对应题目/答案项"]], prompt=True)
        note = clean_ocr_text(row[indexes["备注"]])

        section = sections.setdefault(
            teil,
            {
                "teil": int(teil),
                "title": title,
                "sourceText": "LV2 段落卡片",
                "mode": "paragraph-card",
                "questions": [],
            },
        )

        if paragraph and first_sentence:
            items = [
                {"number": number, "text": prompt}
                for number, prompt in split_answer_items(numbers, item_text)
            ]
            section["questions"].append(
                {
                    "number": paragraph,
                    "text": first_sentence,
                    "german": first_sentence,
                    "chinese": summary,
                    "answer": "",
                    "raw": item_text,
                    "source": "lv2-xlsx",
                    "paragraph": paragraph,
                    "firstSentence": first_sentence,
                    "summary": summary,
                    "items": items,
                    "note": note,
                }
            )

    return [sections[key] for key in sorted(sections, key=lambda value: int(value))]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path, help="Path to LV2 cleaned workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/lv2_questions.js"),
        help="Output JavaScript data file",
    )
    args = parser.parse_args()

    sections = build_sections(args.input)
    payload = json.dumps(sections, ensure_ascii=False, indent=2)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(f"window.LV2_QUESTIONS = {payload};\n", encoding="utf-8")

    question_count = sum(len(section["questions"]) for section in sections)
    print(f"Wrote {len(sections)} LV2 sections and {question_count} questions to {args.output}")


if __name__ == "__main__":
    main()
