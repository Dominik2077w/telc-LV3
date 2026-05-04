#!/usr/bin/env python3
"""Extract LV2 version 2 row-based flashcards from the signal-sentence workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE_SHEET = "LV2信号句总表"
FRONT_COLUMNS = ("Teil", "标题", "段落", "对应题目")
BACK_COLUMNS = (
    "对应段落首句",
    "题型功能",
    "信号所在原文句/句群",
    "对应中文释义",
    "判断逻辑",
    "段落大意",
    "备注",
)


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def row_fields(row: tuple, indexes: dict[str, int], columns: tuple[str, ...]) -> list[dict]:
    fields = []
    for column in columns:
        fields.append({"label": column, "value": clean_text(row[indexes[column]])})
    return fields


def build_sections(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[SOURCE_SHEET]
    rows = list(worksheet.iter_rows(values_only=True))
    header = [clean_text(cell) for cell in rows[0]]
    indexes = {name: index for index, name in enumerate(header)}

    missing = [column for column in (*FRONT_COLUMNS, *BACK_COLUMNS) if column not in indexes]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    sections: dict[str, dict] = {}
    row_number_by_teil: dict[str, int] = {}

    for source_row_number, row in enumerate(rows[1:], start=2):
        teil = clean_text(row[indexes["Teil"]])
        title = clean_text(row[indexes["标题"]])
        paragraph = clean_text(row[indexes["段落"]])
        item = clean_text(row[indexes["对应题目"]])

        if not teil or not title or not paragraph or not item:
            continue

        row_number_by_teil[teil] = row_number_by_teil.get(teil, 0) + 1
        section = sections.setdefault(
            teil,
            {
                "teil": int(teil) if teil.isdigit() else teil,
                "title": title,
                "sourceText": "LV2 版本2 信号句总表",
                "mode": "lv2-v2-card",
                "questions": [],
            },
        )

        section["questions"].append(
            {
                "number": str(row_number_by_teil[teil]),
                "text": f"{paragraph} · {item}",
                "german": item,
                "chinese": clean_text(row[indexes["判断逻辑"]]) or "暂无判断逻辑",
                "answer": "",
                "raw": item,
                "source": "lv2-v2-xlsx",
                "sourceRow": source_row_number,
                "paragraph": paragraph,
                "firstSentence": clean_text(row[indexes["对应段落首句"]]),
                "summary": clean_text(row[indexes["段落大意"]]),
                "frontFields": row_fields(row, indexes, FRONT_COLUMNS),
                "backFields": row_fields(row, indexes, BACK_COLUMNS),
                "items": [{"number": "", "text": item}],
                "note": clean_text(row[indexes["备注"]]),
            }
        )

    return [sections[key] for key in sorted(sections, key=lambda value: int(value) if value.isdigit() else value)]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path, help="Path to LV2 version 2 workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/lv2_v2_questions.js"),
        help="Output JavaScript data file",
    )
    args = parser.parse_args()

    sections = build_sections(args.input)
    payload = json.dumps(sections, ensure_ascii=False, indent=2)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(f"window.LV2_V2_QUESTIONS = {payload};\n", encoding="utf-8")

    question_count = sum(len(section["questions"]) for section in sections)
    print(f"Wrote {len(sections)} LV2 version 2 sections and {question_count} questions to {args.output}")


if __name__ == "__main__":
    main()
